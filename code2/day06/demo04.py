"""
下面，我们来搞定存储数据的另一个基础知识点——Excel文件的读取。
"""
import openpyxl

# 写入的代码：
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'new title'
sheet['A1'] = '漫威宇宙'
rows = [['美国队长', '钢铁侠', '蜘蛛侠', '雷神'], ['是', '漫威', '宇宙', '经典', '人物']]
for i in rows:
    sheet.append(i)
print(rows)
wb.save('Marvel.xlsx')

# 读取的代码：
# 调用openpyxl.load_workbook()函数，打开“Marvel.xlsx”文件。
wb = openpyxl.load_workbook('Marvel.xlsx')
# 获取“Marvel.xlsx”工作薄中名为“new title”的工作表。
sheet = wb['new title']
# sheetnames是用来获取工作薄所有工作表的名字的。如果你不知道工作薄到底有几个工作表，就可以把工作表的名字都打印出来。
sheetname = wb.sheetnames
print(sheetname)
# 把“new title”工作表中A1单元格赋值给A1_cell，再利用单元格value属性，就能打印出A1单元格的值。
A1_cell = sheet['A1']
A1_value = A1_cell.value
print(A1_value)
